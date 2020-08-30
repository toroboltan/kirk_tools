## Imports
import yfinance as yf
import finviz as fvz
import os
from openpyxl import load_workbook

## Parameters & Constants

# File information
tablePath = r'C:\Users\jeron\Google Drive\trading\kirk\2019\active trading'
tableFile = "2020_MoneyManagementSpreadsheet.xlsx"
tableSheet = "trade_log"
tableFileOut = "2020_MoneyManagementSpreadsheet_testb.xlsx"

# Excel info
code_col = 1
date_col = 10
price_col = 11
start_row = 15

# PySimpleGUI Parameters
sectorsOverWindow = (157,28)
sectorsPerfWindow = (190,28)
buttonTradingTable = 'Open Trading Table'

# Years to be selected
year_prefix = 2020

## Functions

# This function returns an Excel workbook
def ReadWorkbook(tablePath, tableFile):
    file_name = os.path.join(tablePath, tableFile)
    print('opening... ' + file_name)
    return load_workbook(filename=file_name)

# This function returns an Excel workbook
def SaveWorkbook(workbook, tablePath, tableFileOut):
    file_name = os.path.join(tablePath, tableFileOut)
    print('saving... ' + file_name)
    workbook.save(filename=file_name)

# This function returns the latest price from finviz
def GetPriceFvz(tkt):
    try:
        price = fvz.get_stock(tkt)['Price']
    except Exception:
        price = 0
        print("price not found in finviz for: " + tkt)
    return price

# This function get trading data for a TKT from 
def GetData(symbol, start='2000-01-01', interval='1d', end=None):
    data = yf.download(symbol, start=start, end=end, interval=interval, auto_adjust=True)
    return data

# This function returns the latest price and based on a 
# flag uses the proper api

def GetPriceTkt(tkt, flag='fv'):
    if flag == 'yf':
        price = GetData(tkt)
    else:
        price = GetPriceFvz(tkt)
    return price

# This function update prices and returns a list of
# open positions

def UpdatePrices():
    workbook = ReadWorkbook(tablePath, tableFile)
    trade_log = workbook[tableSheet]
    set_tkt = set()

    for row in trade_log.iter_rows(min_row=start_row, 
                                   min_col=code_col,
                                   max_col=price_col):
        cell_tkt = row[code_col -1]
        if cell_tkt.value is None:
            break
        else:
            cell_date = row[date_col -1]
            if (cell_date.value.year != year_prefix):
                tkt = cell_tkt.value
                print('getting price for: ' + tkt)
                #getting prices using finviz
                price_tkt = GetPriceTkt(tkt,flag='fv')
                set_tkt.add(tkt)
                cell_price = row[price_col -1]
                cell_price.value = float(price_tkt)
                
    SaveWorkbook(workbook, tablePath, tableFileOut)
    return list(set_tkt)

def main():
    UpdatePrices()

if __name__ == "__main__":
    try:
        main()
    except SystemExit as e:
        print('Error Exception triggered')
