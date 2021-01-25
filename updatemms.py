## Imports
import yfinance as yf
import finviz as fvz
import os
from openpyxl import load_workbook
import kirkconstants
import pandas as pd

# @TODO: parameterize flag values related to type of tool used to get prices

# File information
tablePath = kirkconstants.tablePath
tableFile = kirkconstants.tableFile
tableSheet = kirkconstants.tableSheet
tableFileOut = kirkconstants.tableFileOut

# Excel info
code_col = kirkconstants.code_col
date_col = kirkconstants.date_col
price_col = kirkconstants.price_col
start_row = kirkconstants.start_row

# Years to be selected
year_prefix = kirkconstants.year_prefix

## Functions

def ReadWorkbook(tablePath, tableFile):
    '''
    This function returns an Excel workbook
    '''
    file_name = os.path.join(tablePath, tableFile)
    print('opening... ' + file_name)
    return load_workbook(filename=file_name)

def SaveWorkbook(workbook, tablePath, tableFileOut):
    '''
    This function returns an Excel workbook
    '''
    file_name = os.path.join(tablePath, tableFileOut)
    print('saving... ' + file_name)
    workbook.save(filename=file_name)

def OpenExcelFile(excelPath, excelFile, excelSheet):
    '''
    This function opens an excel file located at a specific
    location and returns a df
    '''
    os.chdir(excelPath)
    return pd.read_excel(excelFile, sheet_name=excelSheet)

def SaveExcelFile(df, excelPath, excelFile, excelSheet):
    '''
    This function writes a df to sheet of a excel file located 
    at a specific location. Does not return anything
    '''
    os.chdir(excelPath)
    df.to_excel(excelFile, sheet_name=excelSheet)
  
def ExportDfToCsv(df, path, filename):
    '''
    This function saves a df on a csv  
    '''
    file_name = os.path.join(path, filename)
    print('saving... ' + file_name)
    df.to_csv(file_name, index=False)

def GetPriceFvz(tkt):
    '''
    This function returns the latest price from finviz
    '''
    # @TODO: parameterize 'Price'
    try:
        price = fvz.get_stock(tkt)['Price']
    except Exception:
        price = 0
        print("price not found in finviz for: " + tkt)
    return price

def GetData(symbol, start='2000-01-01', interval='1d', end=None):
    '''
    This function get trading data for a TKT from 
    '''
    # @TODO: parameterize default values for start and interval
    data = yf.download(symbol, start=start, end=end, interval=interval, auto_adjust=True)
    return data

def GetPriceTkt(tkt, flag='fv'):
    '''
    This function returns the latest price and based on a 
    flag uses the proper api
    '''
    if flag == 'yf':
        price = GetData(tkt)
    else:
        price = GetPriceFvz(tkt)
    return price

def UpdatePrices():
    '''
    This function update prices and returns a list of
    open positions
    '''
    workbook = ReadWorkbook(tablePath, tableFile)
    trade_log = workbook[tableSheet]
    set_tkt = set()

    for row in trade_log.iter_rows(min_row=start_row, 
                                   min_col=code_col,
                                   max_col=price_col):
        cell_tkt = row[code_col -1]
        if cell_tkt.value is None:
            break
        else:
            cell_date = row[date_col -1]
            if (cell_date.value.year == year_prefix):
                tkt = cell_tkt.value
                print('getting price for: ' + tkt)
                #getting prices using finviz
                price_tkt = GetPriceTkt(tkt,flag='fv')
                set_tkt.add(tkt)
                cell_price = row[price_col -1]
                cell_price.value = float(price_tkt)
                
    SaveWorkbook(workbook, tablePath, tableFileOut)
    return list(set_tkt)

def main():
    UpdatePrices()

if __name__ == "__main__":
    try:
        main()
    except SystemExit as e:
        print('Error Exception triggered')
